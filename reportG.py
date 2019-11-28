# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'reportG.ui',
# licensing of 'reportG.ui' applies.
#
# Created: Wed Oct 23 12:19:20 2019
#      by: pyside2-uic  running on PySide2 5.13.1
#
# WARNING! All changes made in this file will be lost!


import openpyxl
from datetime import  date
import random
from report_problem import Problem

from PyQt5 import QtCore, QtGui, QtWidgets


class Ui_Dialog_Client(object):
    def __init__(self, id_client):

        self.id_client = id_client

    def setupUi(self, Dialog_Client):
        Dialog_Client.setObjectName("Dialog_Client")
        Dialog_Client.resize(885, 857)
        self.groupBox = QtWidgets.QGroupBox(Dialog_Client)
        self.groupBox.hide()
        self.groupBox.setEnabled(True)
        self.groupBox.setGeometry(QtCore.QRect(10, 140, 431, 201))
        self.groupBox.setTitle("")
        self.groupBox.setObjectName("groupBox")
        self.label = QtWidgets.QLabel(self.groupBox)
        self.label.setGeometry(QtCore.QRect(20, 30, 71, 21))
        self.label.setObjectName("label")
        self.list_product = QtWidgets.QComboBox(self.groupBox)
        self.list_product.setGeometry(QtCore.QRect(100, 30, 291, 22))
        self.list_product.setObjectName("list_product")
        self.label_2 = QtWidgets.QLabel(self.groupBox)
        self.label_2.setGeometry(QtCore.QRect(20, 80, 51, 21))
        self.label_2.setObjectName("label_2")
        self.lis_problem = QtWidgets.QComboBox(self.groupBox)
        self.lis_problem.setGeometry(QtCore.QRect(100, 80, 291, 22))
        self.lis_problem.setObjectName("lis_problem")
        self.send = QtWidgets.QPushButton(self.groupBox)
        self.send.setGeometry(QtCore.QRect(300, 160, 75, 23))
        self.send.setObjectName("send")
        self.groupBox_2 = QtWidgets.QGroupBox(Dialog_Client)
        self.groupBox_2.hide()
        self.groupBox_2.setGeometry(QtCore.QRect(10, 20, 501, 71))
        self.groupBox_2.setObjectName("groupBox_2")
        self.display = QtWidgets.QPushButton(self.groupBox_2)
        self.display.setGeometry(QtCore.QRect(130, 20, 181, 41))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.display.setFont(font)
        self.display.setObjectName("display")

        self.retranslateUi(Dialog_Client)
        QtCore.QMetaObject.connectSlotsByName(Dialog_Client)

    def retranslateUi(self, Dialog_Client):
        _translate = QtCore.QCoreApplication.translate
        Dialog_Client.setWindowTitle(_translate("Dialog_Client", "Dialog"))
        self.label.setText(_translate("Dialog_Client", "products :"))
        self.label_2.setText(_translate("Dialog_Client", "Problem :"))
        self.send.setText(_translate("Dialog_Client", "send"))
        self.groupBox_2.setTitle(_translate("Dialog_Client", "GroupBox"))
        self.display.setText(_translate("Dialog_Client", "Report Problem"))



    def report_p(self):

        proud_lst = list(self.get_client_product())
        prob_lst = list(self.get_problems_types())
        for i in proud_lst:
            self.list_product.addItem(i[1])
        for j in prob_lst:
            self.lis_problem.addItem(j)
        self.groupBox.show()
        self.list_product.currentTextChanged.connect(self.select_product)
        self.lis_problem.currentTextChanged.connect(self.select_problem)

    def select_product(self):
        choice_product = self.list_product.currentText()
        return choice_product

        # choice_product=self.list_product.currentText()

    def select_problem(self):
        choice_problem = self.lis_problem.currentText()
        return choice_problem

    def get_problems_types(self):
        lsproblems = []
        f = openpyxl.load_workbook('excel_files\\problems_types.xlsx')
        sheet1 = f['types']
        row = sheet1.max_row + 1
        for i in range(2, row):
            lsproblems.append(sheet1.cell(row=i, column=2).value)
        return lsproblems

    def get_client_product(self):
        ans = []
        wb = openpyxl.load_workbook('excel_files\\products.xlsx')
        sheet1 = wb['products']
        for i in range(2, sheet1.max_column + 1):
            if (str(sheet1.cell(row=i, column=3).value) == str(self.id_client)):
                ans.append((sheet1.cell(row=i, column=1).value, sheet1.cell(row=i, column=2).value))
        return ans

    def send_data(self):
        id = self.id_user.text()
        ch_product = self.select_product()
        ch_problem = self.select_problem()
        Problem.add_problem_to_file(ch_product, id, ch_problem)
        print(id, ch_problem, ch_product)




