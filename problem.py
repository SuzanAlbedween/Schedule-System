import openpyxl
import math
from datetime import datetime
velocity = 30

'''
|Problem class
|Files used: problems, problems types, clients
|main functions: 
|1. get_list_id_time 
|   input: type of problems
|   output: list of tiny lists [problem id, problem time]
|2. move_critical_to_regular
|   input: problem id
|   output: problem moved from critical problems sheet to regular problems sheet
|3. delete_Row
|   input: problem id, problem type
|   output: deleting problem from the file
'''


class Problem:

    # written by Suzan added by Abeer
    def GetAllProblemByType(self, type):
        problems = list()
        file = openpyxl.load_workbook('excel_files\\problems.xlsx')
        sheet = file[type]
        rows = sheet.max_row
        for i in range(2, rows + 1):
            detailsproblem = list()
            for j in range(1, 6):
                detailsproblem.append(sheet.cell(row=i, column=j).value)
            problems.append(detailsproblem)
        return problems

    def get_problem_time(self, description):
        problems_types_file = openpyxl.load_workbook('excel_files\\problems_types.xlsx')
        problem_types = problems_types_file['types']
        t_rows = problem_types.max_row
        for j in range(2, t_rows + 1):
            if description == problem_types.cell(row=j, column=2).value:
                return problem_types.cell(row=j, column=3).value
        return 0

    def get_travel_time(self,client_id):
        clients_file = openpyxl.load_workbook('excel_files\\clients.xlsx')
        clients = clients_file['clients']
        c_rows = clients.max_row
        for i in range(2,c_rows+1):
            if int(clients.cell(row=i, column=2).value)== int(client_id):
                location = clients.cell(row=i, column=3).value
                location = [int(i) for i in location.split(',')]
                distance = math.sqrt(location[0] ** 2 + location[1] ** 2)
                return int(distance//velocity)
        return 0

    def get_total_time(self, description, client_id):
        problem_time = self.get_problem_time(description)
        travel_time = self.get_travel_time(client_id)
        return problem_time + travel_time

    def get_list_id_time(self, type_of_problem):
        problems = []
        problems_file = openpyxl.load_workbook('excel_files\\problems.xlsx')
        all_problems = problems_file[type_of_problem]
        p_rows = all_problems.max_row
        for i in range(2, p_rows + 1):
            data = [all_problems.cell(row=i, column=1).value]
            description = all_problems.cell(row=i, column=5).value
            client_id = all_problems.cell(row=i, column=2).value
            data.append(self.get_total_time(description, client_id))
            problems.append(data)
        return problems

    def add_to_regular(self,problem_id, client_id, product_id, report_date, description):
        problems_file = openpyxl.load_workbook('excel_files\\problems.xlsx')
        all_problems = problems_file['regular']
        p_row = all_problems.max_row
        for i in range(2, p_row + 1):
            if all_problems['D'+str(i)].value > report_date:
                all_problems.insert_rows(idx=i, amount=1)
                print(all_problems.cell(row=i, column=1).value)
                all_problems.cell(row=i, column=1).value = problem_id
                all_problems.cell(row=i, column=2).value = client_id
                all_problems.cell(row=i, column=3).value = product_id
                all_problems.cell(row=i, column=4).value = report_date.date()
                all_problems.cell(row=i, column=5).value = description
                print(all_problems.cell(row=i, column=1).value)
                problems_file.save('excel_files\\problems.xlsx')
                break

    def delete_Row(self, problem_id, type_of_problem):
        wb = openpyxl.load_workbook('excel_files\\problems.xlsx')
        sheet = wb[type_of_problem]
        rows = sheet.max_row
        for i in range(2, rows + 1):
            if str(problem_id) == str(sheet.cell(row=i, column=1).value):
                sheet.delete_rows(i, amount=1)
                break
        wb.save('excel_files\\problems.xlsx')

    def move_critical_to_regular(self,problem_id):
        problems_file = openpyxl.load_workbook('excel_files\\problems.xlsx')
        all_problems = problems_file['critical']
        p_row = all_problems.max_row
        for i in range(2,p_row+1):
            if all_problems.cell(row=i,column=1).value == problem_id:
                client_id = all_problems.cell(row=i, column=2).value
                product_id = all_problems.cell(row=i, column=3).value
                report_date = all_problems.cell(row=i,column=4).value
                description = all_problems.cell(row=i,column=5).value
                self.add_to_regular(problem_id, client_id, product_id, report_date, description)
                break
        self.delete_Row(problem_id, 'critical')

