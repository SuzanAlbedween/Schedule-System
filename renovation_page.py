import sys
import openpyxl
from PyQt5.QtWidgets import *
from techns import Techn
from problem import Problem
from PyQt5 import  QtGui
from PyQt5.QtCore import *
#from qtgui import *
from PyQt5.uic.properties import QtWidgets
from scheduling import Scheduling
from renovationGUI import Ui_tech_window

##Techn,Problem,Scheduling


class renovation_App(QMainWindow,Ui_tech_window,Scheduling,Problem):
    def __init__(self):
        QMainWindow.__init__(self)
        Ui_tech_window.__init__(self)
        Scheduling.__init__(self)
        self.setupUi(self)

        self.CreatTabel()
        self.Load_Renovation_PerOneTechn(2)
       # self.table_tech.selectionMode().currentTextChanged.connect(self.Select_Status)
       # self.RenovationForAll()




    def CreatTabel(self):
       Names=["ID Scheduling","Details","Time","Location","Status"]
       for i in range(len(Names)):
           item = QTableWidgetItem()
           item.setText(Names[i])
           self.table_tech.setHorizontalHeaderItem(i, item)


    def Load_Renovation_PerOneTechn(self,index_tech):
        file = openpyxl.load_workbook('excel_files\\scheduling.xlsx')
        sheet = file['sheet1']
        rows = sheet.max_row
        self.id_techlabel.setText(str(sheet.cell(row=1, column=index_tech).value))
        #name=self.Get_TechName(sheet.cell(row=1, column=index_tech).value)
        #print(name)
        #self.nametech_label.setText(name)
        for i in range(rows):
            idproblem =sheet.cell(row=2 + i, column=index_tech).value
            self.table_tech.setItem(i, 0, QTableWidgetItem(str(idproblem)))
            type=self.Type_Of_Problem(idproblem)
           # print(type)
            details=self.Get_ProblemDetails(idproblem,type)
            self.table_tech.setItem(i, 1, QTableWidgetItem(details)) # problemdetails
            #data=self.Getdescription_GetClientID(idproblem,type)
            description=self.Getdescription(idproblem,type)
            Cid=self.GetClientID(idproblem,type)
            totaltime=self.get_total_time(description,Cid)
            self.table_tech.setItem(i, 2, QTableWidgetItem(str(totaltime)))
            location=self.GetLocation(Cid)
            self.table_tech.setItem(i, 3, QTableWidgetItem(location))
            combo = QComboBox()

            if(type=='critical'):

                combo.addItem("None")
                combo.addItem("Done")
                combo.addItem("Convert To Regular")

            else:
                if(type=='regular'):
                 combo.addItem("None")
                 combo.addItem("Done")



            combo.currentTextChanged.connect(self.Select_Status)
            self.table_tech.setCellWidget(i, 4,combo)



           # choice_status=self.table_tech.g

    def Select_Status(self):
        cols=self.table_tech.rowCount()
        for row in range(cols):
            w=self.table_tech.cellWidget(row,4)
            if isinstance(w,  QComboBox):
                #current_value = w.currentText()
                if (str(w.currentText())=="Done"):
                    ppr = self.table_tech.item(row, 0).text()
                    print(ppr)
                    type =str(self.Type_Of_Problem(ppr))
                    print(type)
                    self.delete_Row(ppr, type)
                    break
                if (str(w.currentText()) == "Convert To Regular"):
                    indx=self.table_tech.item(row, 0).text()
                    print(indx)
                    self.move_critical_to_regular(indx)
                    break













#######################run app#######################################
if __name__=="__main__":
    app=QApplication(sys.argv)
    window=renovation_App()
    window.show()
    sys.exit(app.exec_())

