import openpyxl
from datetime import datetime
import random
import string


class Product:

    def id_exists(self, product_id):
        product_file = openpyxl.load_workbook('excel_files\\products.xlsx')
        product_sheet = product_file['products']
        p_row = product_sheet.max_row
        for i in range(2, p_row + 1):
            if product_sheet.cell(row=i, column=2).value == product_id:
                return True
        return False

    def generate_random_id(self):
        product_id = str(random.randrange(1000, 9999)) + random.choice(string.ascii_uppercase)
        if self.id_exists(product_id):
            self.generate_random_id()
        return product_id

    def get_date(self):
        return str(datetime.date(datetime.now()).day) + "/" + str(datetime.date(datetime.now()).month) + "/" + str(
            datetime.date(datetime.now()).year)

    def add_product(self, product_name, client_id):
        products_file = openpyxl.load_workbook('excel_files\\products.xlsx')
        products_sheet = products_file['products']
        p_row = products_sheet.max_row
        products_sheet['A' + str(p_row + 1)] = self.generate_random_id()  # adding id as str
        products_sheet['B' + str(p_row + 1)] = product_name
        products_sheet['C' + str(p_row + 1)] = client_id  # adding id as str
        products_sheet['D' + str(p_row + 1)] = self.get_date()
        products_file.save('excel_files\\products.xlsx')
