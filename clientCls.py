import openpyxl
import random

'''
|Client class
|used files: clients
|main functions:
|1. add_client_to_excel
|   input: client name, client location
|   action: generating random id (5 digit number) and save all of the data in the excel file
|2. testing function of the client:
|   Checking_Name, Testing_Location
'''


class Client:

    def is_id_exist(self, client_id):
        client_file = openpyxl.load_workbook('excel_files\\clients.xlsx')
        client_sheet = client_file['clients']
        row_s = client_sheet.max_row
        for i in range(2, row_s + 1):
            if client_sheet.cell(row=i, column=2).value == client_id:
                return True
        return False

    def client_random_id(self):
        client_id = random.randrange(1000, 99999)
        if self.is_id_exist(client_id) is True:
            self.client_random_id()
        return client_id

    def add_client_to_excel(self, client_name, client_location):
        client_file = openpyxl.load_workbook('excel_files\\clients.xlsx')
        client_sheet = client_file['clients']
        maxrow = client_sheet.max_row
        random_id = (self.client_random_id())
        client_sheet['A' + str(maxrow + 1)] = client_name
        client_sheet['B' + str(maxrow + 1)] = random_id
        client_sheet['C' + str(maxrow + 1)] = client_location
        client_file.save('excel_files\\clients.xlsx')

    def Checking_Length(self, temp):
        if len(temp) < 30:
            return True
        else:
            return False

    def Checking_Name(self, name):
        # The isalpha() method returns True if all characters in the string are alphabets.
        # If not, it returns False.
        if self.Checking_Length(name):
            name = name.split()
            for i in name:
                if not i.isalpha():
                    return False
            return True
        return False

    def Testing_Location(self, location):
        found = location.find(',')
        if found != -1:
            coord = location.split(',')
            # The isnumeric() method returns True if all characters in a string are numeric characters.
            # If not, it returns False.
            if len(coord) == 2:
                if (coord[0].isdecimal() and coord[0] >= '0') and (coord[1].isdecimal() and coord[1] >= '0'):
                    return True
        return False