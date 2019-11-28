import sys
import openpyxl
from PyQt5 import QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5 import  QtGui,QtCore
from PyQt5.QtCore import *
from login_forum import Ui_LoginWindow
from reportG import Ui_Dialog_Client
from clientCls import Client
from OrganizedGUI import Ui_Dialog_Admin
from PyQt5.QtGui import QImage, QPalette, QBrush
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QApplication
import PyQt5.uic as  uic
import os
from PyQt5 import QtGui
from PyQt5.QtWidgets import *
from PyQt5 import  QtGui,QtCore
from PyQt5.QtCore import *
from login_forum import Ui_LoginWindow
from problem import Problem
from product import Product
from techns import Techn
from scheduling import Scheduling
from renovationGUI import Ui_tech_window


class main_login(QMainWindow,Ui_LoginWindow,Problem,Ui_Dialog_Client,Ui_Dialog_Admin,Techn,Product,Client):

    def __init__(self):
        Problem.__init__(self)
        Product.__init__(self)
        Client.__init__(self)
        Techn.__init__(self)
        self.Scheduling = Scheduling()
        Scheduling.__init__(self)
        QMainWindow.__init__(self)
        Ui_LoginWindow.__init__(self)
        self.setupUi(self)

        self.setFixedSize(670, 570)  # sizeofwindow
        # *****************************add image****************

        oImage = QImage("6719.jpg")
        sImage = oImage.scaled(QSize(700, 600))  # resize Image to widgets size
        palette = QPalette()
        palette.setBrush(QPalette.Window, QBrush(sImage))
        self.setPalette(palette)
        self.btn_login.setStyleSheet("background: transparent;")
        self.label.setStyleSheet("color: white;")
        self.btn_login.clicked.connect(self.redirectTo)



    def redirectTo(self):
        username_ = self.user_name.text()
        pwd = self.user_pass.text()
        res_client = self.client_login(username_, pwd)
        res_tech = self.techn_login(username_, pwd)
        res_admin = self.admin_login(username_,pwd)


        if(res_client == 1):

            self.clientWindow = QtWidgets.QMainWindow()
            self.id_client = self.user_pass.text()
            self.ui = Ui_Dialog_Client(self.id_client)
            self.ui.setupUi(self.clientWindow)
            self.hide()
            self.clientWindow.show()

            self.ui.report_p()

            self.ui.send.clicked.connect(self.send_data)

        elif (res_tech == 1):

            QMainWindow.__init__(self)
            self.id_techn = self.user_pass.text()
            self.technWindow = QtWidgets.QMainWindow()
            self.ui = Ui_tech_window(self.id_techn)
            self.ui.setupUi(self.technWindow)
            self.technWindow.show()
            #self.ui.table_tech.hide()
            self.CreatTabel()
            Id=self.ReturnIDTech(username_,pwd)
            index=self.ReturnIndexTech(Id)

            print(index)
            self.Load_Renovation_PerOneTechn(int(index))
            self.ui.table_tech.show()

        elif (res_admin == 1):
            QMainWindow.__init__(self)
            Ui_Dialog_Admin.__init__(self)

            self.adminWindow = QtWidgets.QMainWindow()
            self.ui = Ui_Dialog_Admin()
            self.ui.setupUi_(self.adminWindow)
            self.hide()
            self.adminWindow.show()
            self.off()
            self.ui.add_clientB.clicked.connect(self.addclient)
            self.ui.add_productB.clicked.connect(self.addproduct)
            self.ui.add_techB.clicked.connect(self.addtech)
            self.ui.push_product.clicked.connect(self.pushproduct)
            self.ui.push_client.clicked.connect(self.pushClient)
            self.ui.push_tech.clicked.connect(self.pushtech)






    # Authentication methods
    def client_login(self,name,pwdId):
        clientfile = openpyxl.load_workbook('excel_files\\clients.xlsx')
        clientsheet = clientfile['clients']
        row_s = clientsheet.max_row
        for i in range(2, row_s + 1):
            if (str(clientsheet.cell(row=i, column=2).value) == str(pwdId) and clientsheet.cell(row=i, column=1).value == name):
                return 1
        return 0

    def techn_login(self,username, pwd):
        tech_file = openpyxl.load_workbook('excel_files\\techns.xlsx')
        tc_sheet = tech_file['techns']
        rows = tc_sheet.max_row
        for i in range(2, rows + 1):
            if (str(tc_sheet.cell(row=i, column=3).value) == str(username) and  str(tc_sheet.cell(row=i,column=4).value) == str(pwd)):
                return 1
        return 0

    def admin_login(self, username,pwd):
        if(username=="admin" and pwd=="0123456"):
            return 1
        return 0
#Client page methods:
    def report_p(self):
        id_client=self.user_pass.text()
        proud_lst=list(self.get_client_product())
        prob_lst=list(self.get_problems_types())
        for i in proud_lst:
            self.list_product.addItem(i[1])
        for j in prob_lst:
            self.lis_problem.addItem(j)
        self.ui.groupBox.show()
        self.list_product.currentTextChanged.connect(self.select_product)
        self.lis_problem.currentTextChanged.connect(self.select_problem)


    def select_product(self):
        choice_product=self.list_product.currentText()
        return choice_product
       # choice_product=self.list_product.currentText()
    def select_problem(self):
        choice_problem=self.lis_problem.currentText()
        return choice_problem
    def send_data(self):
        id_client = self.id_client
        ch_product=self.ui.select_product()
        ch_problem=self.ui.select_problem()
        self.add_problem_to_file(ch_product,id_client,ch_problem)
        print(id_client,ch_problem,ch_product)
        #self.add_problem_to_file(choice_product,id_client,choice_problem)
    def get_problems_types(self):
        lsproblems=[]
        f=openpyxl.load_workbook('excel_files\\problems_types.xlsx')
        sheet1 = f['types']
        row = sheet1.max_row + 1
        for i in range(2,row):
            lsproblems.append(sheet1.cell(row=i, column=2).value)
        return lsproblems

    def get_client_product(self):
        ans = []
        wb = openpyxl.load_workbook('excel_files\\products.xlsx')
        sheet1 = wb['products']
        for i in range(2, sheet1.max_column + 1):
            if (str(sheet1.cell(row=i, column=3).value) == str(self.id_client)):
                ans.append((sheet1.cell(row=i, column=1).value, sheet1.cell(row=i, column=2).value))
        return ans

    #Admin page methods:
    def off(self):
        self.ui.groupBox1.hide()
        self.ui.groupBox2.hide()
        self.ui.groupBox3.hide()

    def addclient(self):
        self.off()
        self.ui.groupBox1.show()

    def addproduct(self):
        self.off()
        self.ui.groupBox2.show()
        self.ui.groupBox2.setGeometry(400, 20, 281, 221)

    def addtech(self):
        self.off()
        self.ui.groupBox3.show()
        self.ui.groupBox3.setGeometry(400, 20, 281, 221)

    def pushproduct(self):
        name = self.ui.product_name.text()
        testname=self.Checking_Name(name)
        if(testname==False):
            self.showdialog("The name contains only lower or upper case letters"+"\n"+" and is less than 30 characters long")
        id_client = self.ui.id_pclient.text()
        testid=self.is_id_exist(id_client)
        if(testid==False):
            self.showdialog("The client does not exist")
        self.add_product(name, id_client)

    def pushClient(self):
        n = self.ui.client_name.text()
        tname=self.Checking_Name(n)
        if(tname==False):
            self.showdialog("The name contains only lower or upper case letters" + "\n" + " and is less than 30 characters long")
        lo = self.ui.client_location.text()
        tlocation=self.Testing_Location(lo)
        if(tlocation==False):
            self.showdialog("The location structure contains only two numbers start and end numbers" + "\n" + " and is larger than -1 also in this format" + "\n" + " < number,number>")
        self.add_client_to_excel(n,lo)



    def showdialog(self,meg):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)

        msg.setText(meg)
        retval = msg.exec_()
        print("value of pressed message box button:", retval)


        # self.add_client_to_excel(name,loc)
        # self.add_client_to_excel(name,loc)

    def pushtech(self):
        name = self.ui.tech_name.text()
        ttname=self.Checking_Name(name)
        if(ttname==False):
            self.showdialog("The name contains only lower or upper case letters" + "\n" + " and is less than 30 characters long")
        username = self.ui.tech_username.text()
        tusern=self.Checking_UserName(username)
        if(tusern==False):
            self.showdialog("The username contains only 8  characters"+"\n"+"also the first character is letter"+"\n "+"also without any space")
        password = self.ui.tech_pass.text()
        tpwd=self.Checking_Length(password)
        if(tpwd==False):
            self.showdialog(tpwd)
        self.add_tech(name, username, password)

    #Techn methods:
    def CreatTabel(self):
        Names = ["ID Scheduling", "Details", "Time", "Location", "Status"]
        for i in range(len(Names)):
            item = QTableWidgetItem()
            item.setText(Names[i])
            self.ui.table_tech.setHorizontalHeaderItem(i, item)
            self.ui.table_tech.hide()

    def Load_Renovation_PerOneTechn(self, index_tech):
        file = openpyxl.load_workbook('excel_files\\scheduling.xlsx')
        sheet = file['sheet1']
        rows = sheet.max_row
        self.ui.id_techlabel.setText(str(sheet.cell(row=1, column=int(index_tech)).value))
        # name=self.Get_TechName(sheet.cell(row=1, column=index_tech).value)
        # print(name)
        # self.nametech_label.setText(name)
        for i in range(rows):
            idproblem = sheet.cell(row=2 + i, column=int(index_tech)).value
            self.ui.table_tech.setItem(i, 0, QTableWidgetItem(str(idproblem)))
            type = self.Type_Of_Problem(idproblem)
            # print(type)
            details = self.Scheduling.Get_ProblemDetails(idproblem, type)

            self.ui.table_tech.setItem(i, 1, QTableWidgetItem(details))  # problemdetails

            # data=self.Getdescription_GetClientID(idproblem,type)
            description = self.Getdescription(idproblem, type)
            Cid = self.GetClientID(idproblem, type)
            totaltime = self.get_total_time(description, Cid)
            self.ui.table_tech.setItem(i, 2, QTableWidgetItem(str(totaltime)))
            location = self.Scheduling.GetLocation(Cid)
            self.ui.table_tech.setItem(i, 3, QTableWidgetItem(location))
            combo = QComboBox()

            if (type == 'critical'):

                combo.addItem("None")
                combo.addItem("Done")
                combo.addItem("Convert To Regular")

            else:
                if (type == 'regular'):
                    combo.addItem("None")
                    combo.addItem("Done")

            combo.currentTextChanged.connect(self.Select_Status)

            self.ui.table_tech.setCellWidget(i, 4, combo)
            # self.ui.table_tech.show(index_tech)

        # choice_status=self.table_tech.g

    def Select_Status(self):
        cols = self.table_tech.rowCount()
        for row in range(cols):
            w = self.table_tech.cellWidget(row, 4)
            if isinstance(w, QComboBox):
                # current_value = w.currentText()
                if (str(w.currentText()) == "Done"):
                    ppr = self.table_tech.item(row, 0).text()
                    print(ppr)
                    type = str(self.ype_Of_Problem(ppr))
                    print(type)
                    self.delete_Row(ppr, type)
                    break
                if (str(w.currentText()) == "Convert To Regular"):
                    indx = self.table_tech.item(row, 0).text()
                    print(indx)
                    self.move_critical_to_regular(indx)
                    break


if __name__=="__main__":
    app=QApplication(sys.argv)
    window=main_login()
    window.show()
    sys.exit(app.exec_())
