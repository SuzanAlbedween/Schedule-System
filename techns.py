import random
import openpyxl
'''
|Class Techn
|Used File: techns
|main functions:
|1. add_tech
|   input: name, username, password
|   output: adding tech to excel file 
|2. get_all_ids
|   input: nothing
|   output: updating techID variable which contains all ids
'''

class Techn:

    techID = list()

    def is_tech_exist(self, tech_id):
        tech_file = openpyxl.load_workbook('excel_files\\techns.xlsx')
        tc_sheet = tech_file['techns']
        rows = tc_sheet.max_row
        for i in range(2, rows + 1):
            if tc_sheet.cell(row=i, column=1).value == tech_id:
                return True
        return False

    def gene_tech_id(self):
        tech_id = random.randrange(100000, 999999)
        if self.is_tech_exist(tech_id):
            self.gene_tech_id()
        return tech_id

    def add_tech(self, tech_name, username, pwd):
        wb = openpyxl.load_workbook('excel_files\\techns.xlsx')
        tc_sheet = wb['techns']
        mxrow = tc_sheet.max_row
        id = self.gene_tech_id()
        tc_sheet['A' + str(mxrow + 1)] = id
        tc_sheet['B' + str(mxrow + 1)] = tech_name
        tc_sheet['C' + str(mxrow + 1)] = username
        tc_sheet['D' + str(mxrow + 1)] = pwd
        wb.save('excel_files\\techns.xlsx')

    def get_all_ids(self):
        wb = openpyxl.load_workbook('excel_files\\techns.xlsx')
        tc_sheet = wb['techns']
        mxrow = tc_sheet.max_row
        for i in range(2, mxrow + 1):
            self.techID.append(tc_sheet.cell(row=i, column=1).value)
        wb.save('excel_files\\techns.xlsx')

    def Checking_UserName(self, name):
        if (len(name) != 8):
            return False
        else:
            if name.isspace() == True:
                return False
            else:
                if (name[0].isalpha() == False):
                    return False
                else:
                    return True
    def ReturnIDTech(self, username,pwd):
        tech_file = openpyxl.load_workbook('excel_files\\techns.xlsx')
        tc_sheet = tech_file['techns']
        rows = tc_sheet.max_row
        for i in range(2, rows + 1):
            if (str(tc_sheet.cell(row=i, column=3).value) == str(username) and str( tc_sheet.cell(row=i, column=4).value) == str(pwd)):
                return tc_sheet.cell(row=i, column=1).value



    def ReturnIndexTech(self,id):
        file = openpyxl.load_workbook('excel_files\\scheduling.xlsx')
        sheet = file['sheet1']
        cols = sheet.max_column
        flage = 0
        for i in range(1, cols + 1):
            if (str((sheet.cell(row=1, column=i).value)) == str(id)):
                flage = i
        return flage

