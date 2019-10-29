import openpyxl
import random
class Clent:


    def is_id_exist(self,client_id,sheet):
        clientfile = openpyxl.load_workbook('clients.xlsx')
        clientsheet = clientfile['client']
        clientsheet = clientfile[sheet]
        row_s=clientsheet.max_row
        for i in range(2,row_s+1):
            if(clientsheet.cell(row=i,column=1).value==client_id):
                return True
            return False
    def client_random_id(self,sheet):
        client_id=random.randrange(10000,99999)
        if(self.is_id_exist(client_id) is True):
            if (self.is_id_exist(client_id,sheet) is True):
                self.client_id()


    def add_client_to_excel(self,client_name,client_location):
        print(client_location,client_location)
        client_file=openpyxl.load_workbook('clients.xlsx')
        client_sheet=client_file['client']
        maxrow=client_sheet.max_row
        id=self.client_random_id()
        client_sheet['A'+(maxrow+1)]=id
        client_sheet['B'+(maxrow+1)]=client_name
        client_sheet['C'+(maxrow+1)]=client_location
        client_file.save('clients.xlsx')

