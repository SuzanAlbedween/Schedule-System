import openpyxl
import random

from techns import Techn
from problem import Problem


class Scheduling(Techn, Problem):
    schedule = list()
    techns_ids = list()
    times = list()

    def return_problem_duration(self, name_pm):
        wb = openpyxl.load_workbook('problems_types.xlsx')
        tp_sheet = wb['types']
        rows = tp_sheet.max_row
        for i in range(2, rows + 1):
            if tp_sheet.cell(row=i, column=2).value == name_pm:
                return tp_sheet.cell(row=i, column=3).value

    def Get_ProblemDetails(self, id_problem, type_of_problem):
        st = ""
        file = openpyxl.load_workbook('excel_files\\problems.xlsx')
        sheet = file[type_of_problem]
        rows = sheet.max_row
        for i in range(2, rows + 1):
            if sheet.cell(row=i, column=1).value == id_problem:
                st = st + "Description :"
                des = str(sheet.cell(row=i, column=5).value)
                st = st + des + "\n" + "Type :" + type_of_problem + "\n" + "ID Client :"
                id_client = str(sheet.cell(row=i, column=2).value)
                return st

    def has_time(self, current_techn, time):
        if self.times[current_techn] >= time:
            return True
        else:
            return False

    def list_of_lists(self):
        client_file = openpyxl.load_workbook('excel_files\clients.xlsx')
        client_sheet = client_file['clients']
        maxrw = client_sheet.max_row  # maxrow is the number of tech
        listoflists = []
        for i in range(2, maxrw + 10):
            sublist = [0, 0]
            listoflists.append(sublist)
        return listoflists

    def init_empty_schedule(self, length_of_schedule):
        self.schedule = [[] for i in range(length_of_schedule)]
        return

    def init_times(self, length_of_schedule):
        self.times = [480 for i in range(length_of_schedule)]
        return

    def init_techn_ids(self):
        t = Techn()
        t.get_all_ids()
        self.techns_ids = t.techID

    def get_problems_id_time(self, type_of_problem):
        p = Problem()
        all_problems = p.get_list_id_time(type_of_problem)
        return all_problems

    def update_time(self, current_techn, time):
        self.times[current_techn] -= time

    def problems_schedule(self, type_of_problem):
        all_problems = self.get_problems_id_time(type_of_problem)
        num_of_techns = len(self.techns_ids)
        current_techn = 0
        for problem in all_problems:
            test = 0
            has_time = self.has_time(current_techn, problem[1])
            while not has_time:
                current_techn = (current_techn + 1) % num_of_techns
                test += 1
                if test == num_of_techns:
                    break
                has_time = self.has_time(current_techn, problem[1])
            if has_time:
                self.schedule[current_techn].append(problem)
                self.update_time(current_techn, problem[1])
                current_techn = (current_techn + 1) % num_of_techns

    def main_schedule(self):
        # initialize global vars
        self.init_techn_ids()
        self.init_empty_schedule(len(self.techns_ids))
        self.init_times(len(self.techns_ids))
        self.problems_schedule('critical')
        self.problems_schedule('regular')
        self.SaveScheduling()

    def SaveScheduling(self):
        self.Cleaning('sheet1')  # clear the sheet in scheduling file
        file = openpyxl.load_workbook('excel_files\\scheduling.xlsx')
        sheet = file['sheet1']
        for i in range(len(self.techns_ids)):
            index = sheet.cell(row=1, column=(i + 1))
            var = self.techns_ids[i]
            index.value = var
        for j in range(len(self.schedule)):
            ls = list(self.schedule[j])
            for k in range(len(ls)):
                postion1 = sheet.cell(row=(k + 2), column=(j + 1))
                postion1.value = ls[k][0]
        file.save('excel_files\\scheduling.xlsx')

    def Cleaning(self, NameOfSheet):
        file = openpyxl.load_workbook('excel_files\\scheduling.xlsx')
        sheet = file[NameOfSheet]
        rows = sheet.max_row
        sheet.delete_rows(1, amount=rows)
        file.save('excel_files\\scheduling.xlsx')



