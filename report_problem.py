import openpyxl
from PyQt5.QtCore import QDate, QTime
from datetime import date
import random


class RProblem:

    def get_problems_types(self):
        lsproblems = []
        f = openpyxl.load_workbook('excel_files\\problems_types.xlsx')
        sheet1 = f.get_sheet_by_name('types')
        row = sheet1.max_row + 1
        for i in range(2, row):
            lsproblems.append(sheet1.cell(row=i, column=2).value)
        return lsproblems

    def report(self):
        client_id = input("Enter your id:")
        try_again_if_wrong = 0
        if self.is_client_exist(client_id):
            ans = self.get_client_choice(client_id)
            self.add_problem_to_file([client_id] + ans)
        else:
            try_again_if_wrong += 1
            if try_again_if_wrong == 5:
                return

    def get_client_product(self, client_id):
        ans = []
        wb = openpyxl.load_workbook('excel_files\\products.xlsx')
        sheet1 = wb.get_sheet_by_name('products')
        for i in range(2, sheet1.max_column + 1):
            if sheet1.cell(row=i, column=3).value == client_id:
                ans.append((sheet1.cell(row=i, column=1).value, sheet1.cell(row=i, column=2).value))
        return ans

    def is_client_exist(self, client_id):
        client_file = openpyxl.load_workbook('excel_files\\clients.xlsx')
        client_sheet = client_file.get_sheet_by_name('clients')
        rows = client_sheet.max_row
        for i in range(2, rows + 1):
            if (client_sheet.cell(row=i, column=2).value == client_id):
                return True
        return False

    def get_client_choice(self, id, produc_choice):
        proud_ls = list(self.get_client_product(id))
        ans = []
        product_file = openpyxl.load_workbook('excel_files\\products.xlsx')
        product_sheet1 = product_file.get_sheet_by_name('products')
        for i in proud_ls:
            if (i == produc_choice):
                ans.append((product_sheet1.cell(row=i, column=2).value, product_sheet1.cell(row=i, column=3).value))
                break
        return ans

    def is_probm_exist(self, probm_id, sheet):
        probm_file = openpyxl.load_workbook('excel_files\\problems.xlsx')
        pm_sheet = probm_file[sheet]
        rows = pm_sheet.max_row
        for i in range(2, rows + 1):
            if (pm_sheet.cell(row=i, column=1).value == probm_id):
                return True
        return False

    def gene_problem_id(self, sheet):
        pm_id = random.randrange(1000000, 9999999)
        if (self.is_probm_exist(pm_id, sheet) is True):
            self.gene_problem_id()
        return pm_id

    def return_problem_type(self, problem_name):
        wb = openpyxl.load_workbook('excel_files\\problems_types.xlsx')
        types_sheet = wb['types']
        rows = types_sheet.max_row
        for i in range(2, rows + 1):
            if (types_sheet.cell(row=i, column=2).value == problem_name):
                return types_sheet.cell(row=i, column=1).value

    def return_product_id(self, product_name, client_id):
        wb = openpyxl.load_workbook('excel_files\\products.xlsx')
        product_sheet = wb['products']
        rows = product_sheet.max_row
        print(product_name, client_id, rows)
        for i in range(2, rows + 1):
            print(product_sheet.cell(row=i, column=2).value, product_name,
                  type(product_sheet.cell(row=i, column=3).value), type(client_id))
            if (product_sheet.cell(row=i, column=2).value == product_name and
                    product_sheet.cell(row=i, column=3).value == int(client_id)):
                d = product_sheet.cell(row=i, column=1).value
                return d

    def add_problem_to_file(self, product_name, client_id, descpn):
        typeis = self.return_problem_type(descpn)
        product_id = self.return_product_id(product_name, client_id)
        wb = openpyxl.load_workbook('excel_files\\problems.xlsx')
        p_sheet = wb[typeis]
        id = self.gene_problem_id(typeis)
        mxrow = p_sheet.max_row
        today_date = date.today()
        p_sheet['A' + str(mxrow + 1)] = id
        p_sheet['B' + str(mxrow + 1)] = client_id
        p_sheet['C' + str(mxrow + 1)] = product_id
        p_sheet['D' + str(mxrow + 1)] = today_date
        p_sheet['E' + str(mxrow + 1)] = descpn
        wb.save('excel_files\\problems.xlsx')

    def return_Cproblem_lists(self):
        wb = openpyxl.load_workbook('excel_files\\problems.xlsx')
        p_sheet = wb['critical']
        lists = []
        mxrow = p_sheet.max_row
        print(mxrow)
        for i in range(2, mxrow + 1):
            sublist = []
            sublist.append(p_sheet.cell(row=i, column=1).value)
            sublist.append(p_sheet.cell(row=i, column=2).value)
            sublist.append(p_sheet.cell(row=i, column=5).value)
            lists.append(sublist)
        return lists
